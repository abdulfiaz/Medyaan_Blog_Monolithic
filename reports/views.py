from django.shortcuts import render
from datetime import datetime,timedelta
from django.utils.dateformat import format as date_format
from django.utils import timezone
from rest_framework.views import APIView,status
from rest_framework.response import Response
from events.serializers import *
from events.models import *
from posts.serializers import *
from users.models import PublisherProfile
from users.auth import get_user_roles
from django.conf import settings
from adminapp.iudetail import get_iuobj    
from django.db import transaction
from django.template.loader import render_to_string
from notification.models import TemplateMaster,EventMaster
from adminapp.utils import get_notification
from sdd_blog import settings
from django.http import HttpResponse
from django.utils.timezone import now
from openpyxl import Workbook
from openpyxl.chart import PieChart,BarChart, Reference
from django.db.models import Count
from django.db.models import Sum
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image



# Report genration for event details
class EventReport(APIView):
    def get(self, request):
        domain = request.META.get('HTTP_ORIGIN', settings.APPLICATION_HOST)
        user = request.user
        year = request.GET.get("year")
        month = request.GET.get("month")
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")

        iu_obj = get_iuobj(domain)
        if not iu_obj:
            return Response({"status": "error", "message": "Unauthorized domain"}, status=status.HTTP_401_UNAUTHORIZED)
        
        # user_role = get_user_roles(request)
        # if user_role!= 'manager':
        #     return Response({"status":"error","message":"You are unauthorized to do this action!"},status=status.HTTP_401_UNAUTHORIZED)
        
        event_query = EventDetails.objects.filter(iu_id=iu_obj)

        if year:
            event_query = event_query.filter(event_date__year=year)
        elif month:
            event_query = event_query.filter(event_date__year=year, event_date__month=month)
        elif start_date and end_date:
            event_query = event_query.filter(event_date__range=[start_date, end_date])

        
        data = []
        for event in event_query:
            status = "Upcoming" if event.event_date > now() else "Completed"
            # booked_user_count = EventBookingDetails.objects.filter(event=event).count()
            booked_user_count = (EventBookingDetails.objects.filter(event=event).aggregate(Sum('no_of_tickets')).get('no_of_tickets__sum') or  0)
            total_tickets = booked_user_count + event.event_member_limit
            organizer_profile = UserPersonalProfile.objects.filter(user=event.event_organizer).first()
            organizer_name = f"{organizer_profile.firstname} {organizer_profile.lastname}" 
            data.append([
                event.id,
                organizer_name,
                event.name,
                event.created_at.strftime("%d %b %Y"),
                event.event_date.strftime("%d %b %Y"),
                booked_user_count,
                status,
                event.event_member_limit,
                total_tickets
            ])

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Event Report"

        headers = ["Event ID", "Organizer Name", "Event Name", "Created At", "Event Date", "Booked User Count", "Status", "Available Tickets", "Total Tickets"]

        sheet.merge_cells("A1:I1")
        sheet["A1"] = "Event Report"
        sheet["A1"].font = Font(size=14, bold=True)
        sheet["A1"].alignment = Alignment(horizontal="center")

        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=2, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in data:
            sheet.append(row)

        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            sheet.column_dimensions[column_letter].auto_size = True

        pie = PieChart()
        labels = Reference(sheet, min_col=3, min_row=3, max_row=len(data) + 2)
        data_ref = Reference(sheet, min_col=6, min_row=2, max_row=len(data) + 2)
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "User Count per Event"
        sheet.add_chart(pie, "K2")

        bar_chart = BarChart()
        bar_chart.type = "col"
        bar_chart.style = 10
        bar_chart.title = "Booked User Count vs Total Tickets"
        bar_chart.x_axis.title = "Event"
        bar_chart.y_axis.title = "Tickets"
        
        categories = Reference(sheet, min_col=3, min_row=3, max_row=len(data) + 2) 
        bar_chart_data = Reference(sheet, min_col=6, max_col=9, min_row=2, max_row=len(data) + 2) 
        
        bar_chart.add_data(bar_chart_data, titles_from_data=True)
        bar_chart.set_categories(categories)
        
        bar_chart.x_axis.majorTickMark = "out"
        bar_chart.x_axis.tickLblPos = "low"

        sheet.add_chart(bar_chart, "K20") 

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = "event_report.xlsx"
        response["Content-Disposition"] = f"attachment; filename={filename}"
        workbook.save(response)

        return response

